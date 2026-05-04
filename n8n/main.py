from fastapi import FastAPI, File, UploadFile, Form, Request, Response 
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware    
import traceback   
import os
from fastapi import Form 
import base64  
import io
import matplotlib.pyplot as plt
import inspect 
from claude_routes import router as claude_router
from metavise_routes import router as metavise_router




# ✅ Variável global para armazenar df atual 
df_global = None

# ✅ Variável global para armazenar última análise ou gráfico
ultimo_resultado_texto = ""


# Tentativa segura de importação dos módulos 
try:
    from leitura import ler_arquivo
    from Capabilidade import ANALISES as ANALISES_CAP
    from Exploratoria import ANALISES as ANALISES_EXP 
    from MSA import ANALISES as ANALISES_MSA
    from Kappa import ANALISES as ANALISES_KAPPA
    from Inferencial import ANALISES as ANALISES_INF
    from Preditiva import ANALISES as ANALISES_PRED
    from Controledeprocesso import ANALISES as ANALISES_PROC
    from Analisesdiversas import ANALISES as ANALISES_DIVERSAS
    from graficos import GRAFICOS
    from graficos_interativos import GRAFICOS_INTERATIVOS, CONFIG_GRAFICOS_INTERATIVOS
except ImportError as e:
    print(f"⚠ Erro de importação: {str(e)}")

# Iniciar app
app = FastAPI()
app.include_router(claude_router)
app.include_router(metavise_router)

# Middleware de CORS atualizado para incluir seu novo domínio

app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=(
        r"^https://("
        r".*\.run\.app"                            # Google AI Studio (qualquer subdominio)
        r"|.*\.up\.railway\.app"                   # Railway (qualquer deploy)
        r"|app\.educacaopelotrabalho\.com"         # producao
        r"|aistudio\.google\.com"                  # Google AI Studio principal
        r")$"
        r"|^http://localhost(:\d+)?$"              # localhost para desenvolvimento
    ),
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Responde preflight OPTIONS universalmente
import re as _re_cors
_CORS_REGEX = _re_cors.compile(
    r"^https://("
    r".*\.run\.app"
    r"|.*\.up\.railway\.app"
    r"|app\.educacaopelotrabalho\.com"
    r"|aistudio\.google\.com"
    r")$"
    r"|^http://localhost(:\d+)?$"
)

@app.options("/{path:path}")
async def preflight_handler(request: Request, path: str):
    origin = request.headers.get("origin", "")
    response = Response()
    if origin and _CORS_REGEX.match(origin):
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Access-Control-Allow-Methods"] = "*"
        response.headers["Access-Control-Allow-Headers"] = "*"
        response.headers["Access-Control-Max-Age"] = "86400"
    return response

# Construção do dicionário de análises
ANALISES = {}
ANALISES.update(ANALISES_CAP if 'ANALISES_CAP' in locals() else {})
ANALISES.update(ANALISES_EXP if 'ANALISES_EXP' in locals() else {})
ANALISES.update(ANALISES_INF if 'ANALISES_INF' in locals() else {})
ANALISES.update(ANALISES_MSA if 'ANALISES_MSA' in locals() else {})
ANALISES.update(ANALISES_KAPPA if 'ANALISES_KAPPA' in locals() else {})
ANALISES.update(ANALISES_PRED if 'ANALISES_PRED' in locals() else {})
ANALISES.update(ANALISES_PROC if 'ANALISES_PROC' in locals() else {})
ANALISES.update(ANALISES_DIVERSAS if 'ANALISES_DIVERSAS' in locals() else {})

CONFIG_ANALISES = {
    "Gráfico Sumario": ["df", "coluna_y"],
    "Análise de outliers": ["df", "lista_y"],
    "Correlação de person": ["df", "coluna_y", "lista_x"],
    "Matrix de dispersão": ["df", "coluna_y", "lista_x"],
    "Análise de estabilidade": ["df", "coluna_y"],
    "Análise de limpeza dos dados": ["df"],
    "Análise de cluster": ["df", "lista_x"],
    "Histograma": ["df", "coluna_y", "subgrupo"],
    "Pareto": ["df", "coluna_x","coluna_y", "subgrupo"],
    "Setores (Pizza)": ["df", "coluna_x","coluna_y", "subgrupo"],
    "Barras": ["df", "coluna_x","coluna_y", "subgrupo"],
    "BoxPlot": ["df", "lista_y", "subgrupo"],
    "Dispersão": ["df", "coluna_y", "coluna_x", "subgrupo"],
    "Tendência": ["df", "coluna_y", "Data", "subgrupo"],
    "Bolhas - 3D": ["df", "coluna_y", "coluna_x", "coluna_z"],
    "Superfície - 3D": ["df", "coluna_y", "coluna_x", "coluna_z"],
    "Dispersão 3D": ["df", "coluna_y", "coluna_x", "coluna_z"],
    "Intervalo": ["df", "lista_y", "subgrupo", "field_conf"],
    "Gage R&R": ["df", "coluna_y", "coluna_x", "subgrupo", "field_LIE", "field_LSE"],
    "Concordância de Atributos": ["df", "coluna_y", "coluna_x", "subgrupo", "field", "ordinal"],
    "Vício (Bias)": ["df", "coluna_y", "field", "field_LSE", "field_LIE"],
    "Linearidade": ["df", "coluna_y", "coluna_x", "field_LSE", "field_LIE"],
    "Estabilidade": ["df", "coluna_y", "subgrupo", "field", "field_LSE", "field_LIE"],
    "Método Analítico": ["df", "coluna_y", "coluna_x", "field", "field_LSE", "field_LIE"],
    "1 Sample T": ["df", "coluna_y", "field", "field_conf"],
    "2 Sample T": ["df", "lista_y", "field_conf"],
    "2 Paired Test": ["df", "lista_y", "field_conf"],
    "One way ANOVA": ["df", "lista_y", "subgrupo", "field_conf"],
    "1 Wilcoxon": ["df", "coluna_y", "field", "field_conf"],
    "2 Mann-Whitney": ["df", "lista_y", "field_conf"],
    "2 Wilcoxon Paired": ["df", "lista_y", "field_conf"],
    "Kruskal-Wallis": ["df", "lista_y", "subgrupo", "field_conf"],
    "Friedman Pareado": ["df", "lista_y", "subgrupo", "field_conf"],
    "1 Intervalo de Confianca": ["df", "coluna_y", "field_conf"],
    "1 Intervalo Interquartilico": ["df", "coluna_y"],
    "2 Varianças":["df", "lista_y", "field_conf"],
    "2 Variancas Brown-Forsythe": ["df", "lista_y", "field_conf"],
    "Bartlett": ["df", "lista_y", "subgrupo", "field_conf"],
    "Brown-Forsythe": ["df", "lista_y", "subgrupo", "field_conf"],
    "1 Intervalo de Confianca Variancia": ["df", "coluna_y", "field_conf"],
    "1 Proporcao": ["df", "coluna_x", "field", "field_conf"],
    "2 Proporções": ["df", "coluna_x", "coluna_y"],
    "K Proporcoes": ["df", "lista_y"],
    "Qui-quadrado de Associação": ["df", "coluna_y", "coluna_x"],
    "Qui-quadrado de Ajuste": ["df", "coluna_y", "coluna_x"],
    "Tipo de modelo de regressão": ["df", "coluna_y", "coluna_x"],
    "Regressão Linear": ["df", "coluna_y", "coluna_x"],
    "Regressão Quadrática": ["df", "coluna_y", "coluna_x"],
    "Regressão Cúbica": ["df", "coluna_y", "coluna_x"],
    "Regressão Linear Múltipla": ["df", "coluna_y", "lista_x"],
    "Regressão Binária": ["df", "coluna_y", "lista_x"],
    "Regressão Ordinal": ["df", "coluna_y", "lista_x"],
    "Regressão Nominal": ["df", "coluna_y", "lista_x"],
    "Árvore de Decisão - CART": ["df", "coluna_y", "lista_x"],
    "Random Forest": ["df", "coluna_y", "lista_x"],
    "Série Temporal": ["df", "coluna_y", "Data", "field"],
    "Carta I-MR": ["df", "coluna_y"],
    "Carta X-BarraR": ["df", "coluna_y", "subgrupo"],
    "Carta X-BarraS": ["df", "coluna_y", "subgrupo"], 
    "Carta P": ["df", "coluna_y", "subgrupo"],
    "Carta NP": ["df", "coluna_y", "subgrupo"],
    "Carta C": ["df", "coluna_y"],
    "Carta U": ["df", "coluna_y", "subgrupo"],
    "Carta EWMA": ["df", "coluna_y"],
    "Teste de normalidade": ["df", "coluna_y"],
    "Análise de distribuição estatística": ["df", "coluna_y"],
    "Capabilidade - dados normais": ["df", "coluna_y", "subgrupo", "field_LIE", "field_LSE"],
    "Capabilidade - outras distribuições": ["df", "coluna_y", "subgrupo", "field_dist", "field_LIE", "field_LSE"],
    "Capabilidade - com dados transformados": ["df", "coluna_y", "subgrupo", "field_LIE", "field_LSE"],
    "Capabilidade - com dados discretizados": ["df", "coluna_y", "field_LIE", "field_LSE"],
    "Cálculo de probabilidade": ["df", "coluna_y", "field"],
}


DICIONARIO_TERMOS = {
    "coluna_y": "Uma coluna numérica ou categórica Y selecionada pelo usuário",
    "coluna_x": "Uma coluna numérica ou categórica X selecionada pelo usuário",
    "coluna_z": "Uma coluna numérica Z selecionada pelo usuário",
    "Data": "Coluna de data ou tempo selecionada pelo usuário",
    "lista_y": "Lista de colunas numéricas Y selecionadas pelo usuário; pode conter apenas um valor se subgrupo for acionado",
    "lista_x": "Lista de colunas numéricas ou categóricas X selecionadas pelo usuário; pode conter apenas um valor se subgrupo for acionado",
    "lista_z": "Lista de colunas numéricas Z selecionadas pelo usuário; pode conter apenas um valor se subgrupo for acionado",
    "subgrupo": "Coluna categórica de subgrupos selecionada pelo usuário",
    "field": "Valor de referência ou parâmetro extra (exemplo: valor de H0)",
    "field_conf": "Nível de confiança em porcentagem (exemplo: 95%)",
    "field_dist": "Tipo de distribuição (exemplo: Normal, Weibull)",
    "field_LSE": "Valor do limite superior de engenharia",
    "field_LIE": "Valor do limite inferior de engenharia"
}

# ✅ Variável global para armazenar df atual
df_global = None

# ✅ Variável global para armazenar última análise ou gráfico
ultimo_resultado_texto = ""

@app.post("/analise")
async def analisar(
    request: Request,
    arquivo: UploadFile = File(None),
    ferramenta: str = Form(None),
    grafico: str = Form(None),
    aba: str = Form(None),
    coluna_y: str = Form(None),
    coluna_x: str = Form(None),
    coluna_z: str = Form(None),
    lista_y: str = Form(None),
    lista_x: str = Form(None),
    subgrupo: str = Form(None),
    field: str = Form(None),
    field_conf: str = Form(None),
    field_dist: str = Form(None),
    field_LSE: str = Form(None),
    field_LIE: str = Form(None),
    Data: str = Form(None),
    pergunta: str = Form(None)
):
    try:
        if not arquivo:
            return JSONResponse({"erro": "Nenhum arquivo recebido.", "analise": ""}, status_code=400)

        df = await ler_arquivo(arquivo, aba)
        global df_global
        df_global = df

        if df is None or df.empty:
            return JSONResponse({"erro": "Arquivo vazio ou aba inválida.", "analise": ""}, status_code=400)

        lista_y_processada = [x.strip() for x in lista_y.split(",")] if lista_y else []
        lista_x_processada = [x.strip() for x in lista_x.split(",")] if lista_x else []
        lista_z = [coluna_z.strip()] if coluna_z else []
        subgrupo_val = subgrupo.strip() if subgrupo else None

        resultado_texto = ""
        imagem_analise_base64 = None
        imagem_grafico_isolado_base64 = None
        info_grafico = None
        resposta_ia = None

        global ultimo_resultado_texto

        if ferramenta:
            funcao = ANALISES.get(ferramenta.strip())
            if not funcao:
                return JSONResponse({"erro": f"Análise {ferramenta} desconhecida.", "analise": ""}, status_code=400)

            disponiveis = {
                "df": df,
                "coluna_y": coluna_y.strip() if coluna_y else None,
                "coluna_x": coluna_x.strip() if coluna_x else None,
                "coluna_z": coluna_z.strip() if coluna_z else None,
                "Data": Data,
                "lista_y": lista_y_processada,
                "lista_x": lista_x_processada,
                "lista_z": lista_z,
                "subgrupo": subgrupo_val,
                "field": field,
                "field_conf": field_conf,
                "field_dist": field_dist,
                "field_LSE": field_LSE,
                "field_LIE": field_LIE
            }

            permitidos = CONFIG_ANALISES.get(ferramenta.strip(), ["df"])
            args_to_pass = {k: disponiveis[k] for k in permitidos if k in disponiveis}
            resultado_texto, imagem_analise_base64 = funcao(**args_to_pass)
            ultimo_resultado_texto = resultado_texto

        if grafico:
            funcao_grafico = GRAFICOS.get(grafico.strip())
            if not funcao_grafico:
                return JSONResponse({"erro": f"Gráfico {grafico} não encontrado.", "analise": ""}, status_code=400)

            disponiveis = {
                "df": df,
                "coluna_y": coluna_y.strip() if coluna_y else None,
                "coluna_x": coluna_x.strip() if coluna_x else None,
                "coluna_z": coluna_z.strip() if coluna_z else None,
                "Data": Data,
                "lista_y": lista_y_processada,
                "lista_x": lista_x_processada,
                "lista_z": lista_z,
                "subgrupo": subgrupo_val,
                "field": field,
                "field_conf": field_conf,
                "field_dist": field_dist,
                "field_LSE": field_LSE,
                "field_LIE": field_LIE
            }

            permitidos = CONFIG_ANALISES.get(grafico.strip(), ["df", "coluna_y"])
            args_to_pass = {k: disponiveis[k] for k in permitidos if k in disponiveis}

            import inspect
            params_aceitos = inspect.signature(funcao_grafico).parameters
            args_filtrados = {k: v for k, v in args_to_pass.items() if k in params_aceitos}

            # 🔧 Recebe mensagem, imagem e info_grafico
            mensagem, imagem_grafico_isolado_base64, info_grafico = funcao_grafico(**args_filtrados)
            ultimo_resultado_texto = f"Gráfico gerado: {grafico}"

        return JSONResponse({
            "analise": resultado_texto,
            "grafico_base64": imagem_analise_base64,
            "grafico_isolado_base64": imagem_grafico_isolado_base64,
            "info_grafico": info_grafico,
            "resposta_ia": resposta_ia
        })

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        return JSONResponse({
            "erro": "Erro interno ao processar a análise.",
            "analise": "",
            "detalhe": str(e),
            "traceback": tb
        }, status_code=500)



    # ✅ SE TIVER PERGUNTA: chama o agente IA sobre a análise ou gráfico
    if ferramenta:
        funcao = ANALISES.get(ferramenta.strip())
        if not funcao:
            return JSONResponse({"erro": f"Análise {ferramenta} desconhecida."}, status_code=400)

        disponiveis = {
            "df": df,
            "coluna_y": coluna_y.strip() if coluna_y else None,
            "coluna_x": coluna_x.strip() if coluna_x else None,
            "coluna_z": coluna_z.strip() if coluna_z else None,
            "Data": Data,
            "lista_y": lista_y_processada,
            "lista_x": lista_x_processada,
            "lista_z": lista_z,
            "subgrupo": subgrupo_val,
            "field": field,
            "field_conf": field_conf,
            "field_dist": field_dist,
            "field_LSE": field_LSE,
            "field_LIE": field_LIE
        }

        permitidos = CONFIG_ANALISES.get(ferramenta.strip(), ["df"])
        args_to_pass = {k: disponiveis[k] for k in permitidos if k in disponiveis}
        resultado_texto, imagem_analise_base64 = funcao(**args_to_pass)

        ultimo_resultado_texto = resultado_texto  # ✅ atualiza global

        if pergunta:
            from agente import perguntar_ia
            resposta_ia = perguntar_ia(pergunta, ultimo_resultado_texto)

    if grafico:
        funcao_grafico = GRAFICOS.get(grafico.strip())
        if not funcao_grafico:
            return JSONResponse({"erro": f"Gráfico {grafico} não encontrado."}, status_code=400)

        disponiveis = {
            "df": df,
            "coluna_y": coluna_y.strip() if coluna_y else None,
            "coluna_x": coluna_x.strip() if coluna_x else None,
            "coluna_z": coluna_z.strip() if coluna_z else None,
            "Data": Data,
            "lista_y": lista_y_processada,
            "lista_x": lista_x_processada,
            "lista_z": lista_z,
            "subgrupo": subgrupo_val,
            "field": field,
            "field_conf": field_conf,
            "field_dist": field_dist,
            "field_LSE": field_LSE,
            "field_LIE": field_LIE
        }

        permitidos = CONFIG_ANALISES.get(grafico.strip(), ["df", "coluna_y"])
        args_to_pass = {k: disponiveis[k] for k in permitidos if k in disponiveis}

        import inspect
        params_aceitos = inspect.signature(funcao_grafico).parameters
        args_filtrados = {k: v for k, v in args_to_pass.items() if k in params_aceitos}

        imagem_grafico_isolado_base64 = funcao_grafico(**args_filtrados)

        ultimo_resultado_texto = f"Gráfico gerado: {grafico}"  # ✅ atualiza global com descrição do gráfico

        if pergunta:
            from agente import perguntar_ia
            resposta_ia = perguntar_ia(pergunta, ultimo_resultado_texto)




@app.post("/personalizar-grafico")
async def personalizar_grafico(
    request: Request,
    arquivo: UploadFile = File(None),
    aba: str = Form(None),
    grafico: str = Form(None),
    coluna_y: str = Form(None),
    coluna_x: str = Form(None),
    coluna_z: str = Form(None),
    subgrupo: str = Form(None),
    field: str = Form(None),
    field_conf: str = Form(None),
    field_dist: str = Form(None),
    field_LSE: str = Form(None),
    field_LIE: str = Form(None),
    Data: str = Form(None),
    lista_y: str = Form(None),
    lista_x: str = Form(None),
    cor: str = Form(""),
    titulo_x: str = Form(""),
    titulo_y: str = Form(""),
    titulo_grafico: str = Form(""),
    tamanho_fonte: str = Form(""),
    inclinacao_x: str = Form(""),
    inclinacao_y: str = Form(""),
    espessura: str = Form("")
):
    try:
        global df_global
        from graficos import GRAFICOS
        import inspect

        print("\n====================== INÍCIO DEBUG /PERSONALIZAR-GRAFICO ======================")
        print("🎨 Gráfico solicitado:", grafico)
        print("➡ coluna_y:", coluna_y)
        print("➡ cor:", cor)
        print("➡ titulo_x:", titulo_x)
        print("➡ titulo_y:", titulo_y)
        print("➡ titulo_grafico:", titulo_grafico)
        print("➡ tamanho_fonte:", tamanho_fonte)
        print("➡ inclinacao_x:", inclinacao_x)
        print("➡ inclinacao_y:", inclinacao_y)
        print("➡ espessura:", espessura)
        print("======================= FIM DEBUG /PERSONALIZAR-GRAFICO ==========================\n")

        df = df_global
        if df is None or df.empty:
            return JSONResponse({"erro": "Nenhum DataFrame carregado. Gere o gráfico primeiro."}, status_code=400)

        lista_y_processada = [x.strip() for x in lista_y.split(",")] if lista_y else []
        lista_x_processada = [x.strip() for x in lista_x.split(",")] if lista_x else []

        funcao_grafico = GRAFICOS.get(grafico.strip())
        if not funcao_grafico:
            return JSONResponse({"erro": f"Gráfico {grafico} não encontrado."}, status_code=400)

        params_aceitos = inspect.signature(funcao_grafico).parameters

        args_to_pass = {k: v for k, v in {
            "df": df,
            "coluna_y": coluna_y,
            "coluna_x": coluna_x,
            "coluna_z": coluna_z,
            "subgrupo": subgrupo,
            "field": field,
            "field_conf": field_conf,
            "field_dist": field_dist,
            "field_LSE": field_LSE,
            "field_LIE": field_LIE,
            "Data": Data,
            "lista_y": lista_y_processada,
            "lista_x": lista_x_processada,
            "cor": cor or "",
            "titulo_x": titulo_x or "",
            "titulo_y": titulo_y or "",
            "titulo_grafico": titulo_grafico or "",
            "tamanho_fonte": tamanho_fonte or "",
            "inclinacao_x": inclinacao_x or "",
            "inclinacao_y": inclinacao_y or "",
            "espessura": espessura or ""
        }.items() if k in params_aceitos}

        imagem_grafico_isolado_base64, info_grafico = funcao_grafico(**args_to_pass)

        return {
            "grafico_isolado_base64": imagem_grafico_isolado_base64,
            "info_grafico": info_grafico  # ✅ Retorna info_grafico com os dados salvos
        }

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        return JSONResponse({
            "erro": "Erro interno ao personalizar gráfico.",
            "detalhe": str(e),
            "traceback": tb
        }, status_code=500)




@app.post("/pergunta")
async def pergunta(request: Request, pergunta: str = Form(...), tipo: str = Form(...)):
    try:
        from agente import perguntar_ia
        global ultimo_resultado_texto

        if tipo == "analise":
            texto_base = ultimo_resultado_texto or "Nenhuma análise encontrada."
        elif tipo == "grafico":
            texto_base = "Último gráfico gerado no sistema."
        else:
            return JSONResponse({"erro": "Tipo de pergunta inválido."}, status_code=400)

        # 🔧 Novo: cria prompt completo unindo análise + pergunta
        prompt_completo = f"""
        Esta é a última análise ou gráfico gerado pelo sistema:

        {texto_base}

        Agora responda a seguinte pergunta do aluno, considerando as informações acima:

        {pergunta}
        """

        resposta = perguntar_ia(pergunta, prompt_completo)

        return {"resposta": resposta}

    except Exception as e:
        tb = traceback.format_exc()
        return JSONResponse({
            "erro": "Erro interno ao processar a pergunta.",
            "detalhe": str(e),
            "traceback": tb
        }, status_code=500)


# ════════════════════════════════════════════════════════════════════
# ROTAS V2 — copia literal das rotas originais para uso pelo Copilot
# As rotas /analise, /personalizar-grafico e /pergunta acima continuam
# intactas servindo o html-main.
# ════════════════════════════════════════════════════════════════════

@app.post("/v2/analise")
async def analisar_v2(
    request: Request,
    arquivo: UploadFile = File(None),
    ferramenta: str = Form(None),
    grafico: str = Form(None),
    aba: str = Form(None),
    coluna_y: str = Form(None),
    coluna_x: str = Form(None),
    coluna_z: str = Form(None),
    lista_y: str = Form(None),
    lista_x: str = Form(None),
    subgrupo: str = Form(None),
    field: str = Form(None),
    field_conf: str = Form(None),
    field_dist: str = Form(None),
    field_LSE: str = Form(None),
    field_LIE: str = Form(None),
    Data: str = Form(None),
    ordinal: str = Form(None),
    pergunta: str = Form(None)
):
    try:
        if not arquivo:
            return JSONResponse({"erro": "Nenhum arquivo recebido.", "analise": ""}, status_code=400)

        df = await ler_arquivo(arquivo, aba)
        global df_global
        df_global = df

        if df is None or df.empty:
            return JSONResponse({"erro": "Arquivo vazio ou aba inválida.", "analise": ""}, status_code=400)

        lista_y_processada = [x.strip() for x in lista_y.split(",")] if lista_y else []
        lista_x_processada = [x.strip() for x in lista_x.split(",")] if lista_x else []
        lista_z = [coluna_z.strip()] if coluna_z else []
        subgrupo_val = subgrupo.strip() if subgrupo else None

        resultado_texto = ""
        imagem_analise_base64 = None
        imagem_grafico_isolado_base64 = None
        info_grafico = None
        resposta_ia = None

        global ultimo_resultado_texto

        if ferramenta:
            funcao = ANALISES.get(ferramenta.strip())
            if not funcao:
                return JSONResponse({"erro": f"Análise {ferramenta} desconhecida.", "analise": ""}, status_code=400)

            disponiveis = {
                "df": df,
                "coluna_y": coluna_y.strip() if coluna_y else None,
                "coluna_x": coluna_x.strip() if coluna_x else None,
                "coluna_z": coluna_z.strip() if coluna_z else None,
                "Data": Data,
                "lista_y": lista_y_processada,
                "lista_x": lista_x_processada,
                "lista_z": lista_z,
                "subgrupo": subgrupo_val,
                "field": field,
                "field_conf": field_conf,
                "field_dist": field_dist,
                "field_LSE": field_LSE,
                "field_LIE": field_LIE,
                "ordinal": (ordinal == "true")
            }

            permitidos = CONFIG_ANALISES.get(ferramenta.strip(), ["df"])
            args_to_pass = {k: disponiveis[k] for k in permitidos if k in disponiveis}
            resultado_texto, imagem_analise_base64 = funcao(**args_to_pass)
            ultimo_resultado_texto = resultado_texto

        if grafico:
            funcao_grafico = GRAFICOS.get(grafico.strip())
            if not funcao_grafico:
                return JSONResponse({"erro": f"Gráfico {grafico} não encontrado.", "analise": ""}, status_code=400)

            disponiveis = {
                "df": df,
                "coluna_y": coluna_y.strip() if coluna_y else None,
                "coluna_x": coluna_x.strip() if coluna_x else None,
                "coluna_z": coluna_z.strip() if coluna_z else None,
                "Data": Data,
                "lista_y": lista_y_processada,
                "lista_x": lista_x_processada,
                "lista_z": lista_z,
                "subgrupo": subgrupo_val,
                "field": field,
                "field_conf": field_conf,
                "field_dist": field_dist,
                "field_LSE": field_LSE,
                "field_LIE": field_LIE,
                "ordinal": (ordinal == "true")
            }

            permitidos = CONFIG_ANALISES.get(grafico.strip(), ["df", "coluna_y"])
            args_to_pass = {k: disponiveis[k] for k in permitidos if k in disponiveis}

            import inspect
            params_aceitos = inspect.signature(funcao_grafico).parameters
            args_filtrados = {k: v for k, v in args_to_pass.items() if k in params_aceitos}

            # 🔧 Recebe mensagem, imagem e info_grafico
            mensagem, imagem_grafico_isolado_base64, info_grafico = funcao_grafico(**args_filtrados)
            ultimo_resultado_texto = f"Gráfico gerado: {grafico}"

        return JSONResponse({
            "analise": resultado_texto,
            "grafico_base64": imagem_analise_base64,
            "grafico_isolado_base64": imagem_grafico_isolado_base64,
            "info_grafico": info_grafico,
            "resposta_ia": resposta_ia
        })

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        return JSONResponse({
            "erro": "Erro interno ao processar a análise.",
            "analise": "",
            "detalhe": str(e),
            "traceback": tb
        }, status_code=500)


@app.post("/v2/personalizar-grafico")
async def personalizar_grafico_v2(
    request: Request,
    arquivo: UploadFile = File(None),
    aba: str = Form(None),
    grafico: str = Form(None),
    coluna_y: str = Form(None),
    coluna_x: str = Form(None),
    coluna_z: str = Form(None),
    subgrupo: str = Form(None),
    field: str = Form(None),
    field_conf: str = Form(None),
    field_dist: str = Form(None),
    field_LSE: str = Form(None),
    field_LIE: str = Form(None),
    Data: str = Form(None),
    lista_y: str = Form(None),
    lista_x: str = Form(None),
    cor: str = Form(""),
    titulo_x: str = Form(""),
    titulo_y: str = Form(""),
    titulo_grafico: str = Form(""),
    tamanho_fonte: str = Form(""),
    inclinacao_x: str = Form(""),
    inclinacao_y: str = Form(""),
    espessura: str = Form(""),
):
    try:
        global df_global
        from graficos import GRAFICOS
        import inspect

        print("\n====================== INÍCIO DEBUG /V2/PERSONALIZAR-GRAFICO ======================")
        print("🎨 Gráfico solicitado:", grafico)
        print("➡ coluna_y:", coluna_y)
        print("➡ cor:", cor)
        print("➡ titulo_x:", titulo_x)
        print("➡ titulo_y:", titulo_y)
        print("➡ titulo_grafico:", titulo_grafico)
        print("➡ tamanho_fonte:", tamanho_fonte)
        print("➡ inclinacao_x:", inclinacao_x)
        print("➡ inclinacao_y:", inclinacao_y)
        print("➡ espessura:", espessura)
        print("======================= FIM DEBUG /V2/PERSONALIZAR-GRAFICO ==========================\n")

        df = df_global
        if df is None or df.empty:
            return JSONResponse({"erro": "Nenhum DataFrame carregado. Gere o gráfico primeiro."}, status_code=400)

        lista_y_processada = [x.strip() for x in lista_y.split(",")] if lista_y else []
        lista_x_processada = [x.strip() for x in lista_x.split(",")] if lista_x else []

        funcao_grafico = GRAFICOS.get(grafico.strip())
        if not funcao_grafico:
            return JSONResponse({"erro": f"Gráfico {grafico} não encontrado."}, status_code=400)

        params_aceitos = inspect.signature(funcao_grafico).parameters

        args_to_pass = {k: v for k, v in {
            "df": df,
            "coluna_y": coluna_y,
            "coluna_x": coluna_x,
            "coluna_z": coluna_z,
            "subgrupo": subgrupo,
            "field": field,
            "field_conf": field_conf,
            "field_dist": field_dist,
            "field_LSE": field_LSE,
            "field_LIE": field_LIE,
            "Data": Data,
            "lista_y": lista_y_processada,
            "lista_x": lista_x_processada,
            "cor": cor or "",
            "titulo_x": titulo_x or "",
            "titulo_y": titulo_y or "",
            "titulo_grafico": titulo_grafico or "",
            "tamanho_fonte": tamanho_fonte or "",
            "inclinacao_x": inclinacao_x or "",
            "inclinacao_y": inclinacao_y or "",
            "espessura": espessura or ""
        }.items() if k in params_aceitos}

        imagem_grafico_isolado_base64, info_grafico = funcao_grafico(**args_to_pass)

        return {
            "grafico_isolado_base64": imagem_grafico_isolado_base64,
            "info_grafico": info_grafico  # ✅ Retorna info_grafico com os dados salvos
        }

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        return JSONResponse({
            "erro": "Erro interno ao personalizar gráfico.",
            "detalhe": str(e),
            "traceback": tb
        }, status_code=500)


@app.post("/v2/pergunta")
async def pergunta_v2(request: Request, pergunta: str = Form(...), tipo: str = Form(...)):
    try:
        from agente import perguntar_ia
        global ultimo_resultado_texto

        if tipo == "analise":
            texto_base = ultimo_resultado_texto or "Nenhuma análise encontrada."
        elif tipo == "grafico":
            texto_base = "Último gráfico gerado no sistema."
        else:
            return JSONResponse({"erro": "Tipo de pergunta inválido."}, status_code=400)

        # 🔧 Novo: cria prompt completo unindo análise + pergunta
        prompt_completo = f"""
        Esta é a última análise ou gráfico gerado pelo sistema:

        {texto_base}

        Agora responda a seguinte pergunta do aluno, considerando as informações acima:

        {pergunta}
        """

        resposta = perguntar_ia(pergunta, prompt_completo)

        return {"resposta": resposta}

    except Exception as e:
        tb = traceback.format_exc()
        return JSONResponse({
            "erro": "Erro interno ao processar a pergunta.",
            "detalhe": str(e),
            "traceback": tb
        }, status_code=500)


# ════════════════════════════════════════════════════════════════════
# ENDPOINT — GERAR PLANILHA DE COLETA GAGE R&R
# Adicionar este bloco ao main.py (em qualquer lugar após as outras
# rotas /v2/...). Não toque em nada já existente.
# ════════════════════════════════════════════════════════════════════
 
@app.post("/v2/gerar-planilha-gage-rr")
async def gerar_planilha_gage_rr(
    request: Request,
    n_pecas: int = Form(...),
    n_operadores: int = Form(...),
    n_replicas: int = Form(...),
    ordem: str = Form("aleatorio"),  # "aleatorio" ou "sequencial"
):
    """
    Gera uma planilha Excel de coleta para estudo Gage R&R Cruzado.
    Saída: arquivo .xlsx com 3 colunas (Peça, Operador, Medição).
    A coluna Medição vem em branco para o usuário preencher no chão de fábrica.
    """
    try:
        import io
        import random
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from fastapi.responses import StreamingResponse
 
        # ============== VALIDAÇÕES ==============
        if n_pecas < 2 or n_pecas > 100:
            return JSONResponse({"erro": "Número de peças deve estar entre 2 e 100."}, status_code=400)
        if n_operadores < 2 or n_operadores > 50:
            return JSONResponse({"erro": "Número de operadores deve estar entre 2 e 50."}, status_code=400)
        if n_replicas < 2 or n_replicas > 50:
            return JSONResponse({"erro": "Número de réplicas deve estar entre 2 e 50."}, status_code=400)
 
        # ============== GERAR NOMES PADRÃO ==============
        nomes_pecas = [f"Peça {str(i+1).zfill(2)}" for i in range(n_pecas)]
        nomes_operadores = [f"Operador {chr(65+i)}" for i in range(n_operadores)]  # A, B, C, ...
 
        # ============== MONTAR LINHAS ==============
        # Para cada réplica, cada operador mede todas as peças.
        # Total de linhas = n_pecas × n_operadores × n_replicas
        linhas = []
        for replica in range(1, n_replicas + 1):
            for operador in nomes_operadores:
                pecas_desta_rodada = nomes_pecas.copy()
                if ordem == "aleatorio":
                    random.shuffle(pecas_desta_rodada)
                for peca in pecas_desta_rodada:
                    linhas.append({
                        "Peça": peca,
                        "Operador": operador,
                        "Medição": ""  # em branco para preencher
                    })
 
        # ============== CRIAR EXCEL ==============
        wb = Workbook()
        ws = wb.active
        ws.title = "Gage R&R"
 
        # Cabeçalho com formatação
        headers = ["Peça", "Operador", "Medição"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
 
        # Dados
        for row_idx, linha in enumerate(linhas, start=2):
            ws.cell(row=row_idx, column=1, value=linha["Peça"])
            ws.cell(row=row_idx, column=2, value=linha["Operador"])
            ws.cell(row=row_idx, column=3, value=linha["Medição"])
 
        # Larguras de coluna
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
 
        # Aba de instruções
        ws_inst = wb.create_sheet("Instruções")
        instrucoes = [
            ["📋 PLANILHA DE COLETA — GAGE R&R"],
            [""],
            [f"Configuração: {n_pecas} peças × {n_operadores} operadores × {n_replicas} réplicas"],
            [f"Total de medições: {len(linhas)}"],
            [f"Ordem das medições: {'Aleatória' if ordem == 'aleatorio' else 'Sequencial'}"],
            [""],
            ["INSTRUÇÕES DE PREENCHIMENTO:"],
            ["1. Não altere as colunas 'Peça' e 'Operador' — siga a ordem proposta."],
            ["2. Cada operador mede cada peça o número de réplicas indicado, sem ver"],
            ["   medições anteriores."],
            ["3. Preencha apenas a coluna 'Medição' com o valor obtido."],
            ["4. Após preencher tudo, faça upload desta planilha de volta no sistema"],
            ["   no modo 'Analisar planilha preenchida'."],
            [""],
            ["BOAS PRÁTICAS:"],
            ["• Use peças que cubram a faixa de variação do processo."],
            ["• Operadores devem ser representativos dos que normalmente medem."],
            ["• Não revele aos operadores qual peça eles estão medindo."],
        ]
        for row_idx, linha in enumerate(instrucoes, start=1):
            cell = ws_inst.cell(row=row_idx, column=1, value=linha[0] if linha else "")
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="1F2937")
            elif "INSTRUÇÕES" in (linha[0] if linha else "") or "BOAS PRÁTICAS" in (linha[0] if linha else ""):
                cell.font = Font(bold=True, size=11, color="2563EB")
        ws_inst.column_dimensions["A"].width = 80
 
        # Salvar em memória
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
 
        nome_arquivo = f"gage_rr_{n_pecas}p_{n_operadores}op_{n_replicas}r.xlsx"
 
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"}
        )
 
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        return JSONResponse({
            "erro": "Erro interno ao gerar a planilha.",
            "detalhe": str(e),
            "traceback": tb
        }, status_code=500)


@app.post("/v2/grafico-interativo")
async def grafico_interativo(
    request: Request,
    arquivo: UploadFile = File(None),
    aba: str = Form(None),
    grafico: str = Form(None),
    coluna_y: str = Form(None),
    coluna_x: str = Form(None),
    coluna_z: str = Form(None),
    lista_y: str = Form(None),
    lista_x: str = Form(None),
    subgrupo: str = Form(None),
    field: str = Form(None),
    Data: str = Form(None),
):
    """
    Endpoint enxuto que delega a renderização para o módulo graficos_interativos.
    Retorna JSON estruturado para o Plotly.js renderizar no front.
    """
    try:
        if arquivo is None:
            return JSONResponse({"erro": "Envie um arquivo Excel."}, status_code=400)

        if not grafico:
            return JSONResponse({"erro": "Informe o tipo de gráfico."}, status_code=400)

        if grafico not in GRAFICOS_INTERATIVOS:
            disponiveis = ", ".join(GRAFICOS_INTERATIVOS.keys())
            return JSONResponse(
                {"erro": f"Gráfico '{grafico}' não suportado em modo interativo. Disponíveis: {disponiveis}"},
                status_code=400
            )

        import pandas as pd
        import io
        file_bytes = await arquivo.read()
        if arquivo.filename.endswith(".xlsx") or arquivo.filename.endswith(".xlsm"):
            df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl", sheet_name=aba)
        elif arquivo.filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(file_bytes))
        else:
            return JSONResponse({"erro": "Formato não suportado. Use .xlsx, .xlsm ou .csv."}, status_code=400)

        lista_y_proc = [v.strip() for v in lista_y.split(",")] if lista_y else None
        lista_x_proc = [v.strip() for v in lista_x.split(",")] if lista_x else None

        disponiveis = {
            "df": df,
            "coluna_y": coluna_y.strip() if coluna_y else None,
            "coluna_x": coluna_x.strip() if coluna_x else None,
            "coluna_z": coluna_z.strip() if coluna_z else None,
            "subgrupo": subgrupo.strip() if subgrupo else None,
            "lista_y": lista_y_proc,
            "lista_x": lista_x_proc,
            "field": field,
            "Data": Data,
        }

        parametros_necessarios = CONFIG_GRAFICOS_INTERATIVOS.get(grafico, ["df"])
        argumentos = {p: disponiveis.get(p) for p in parametros_necessarios}

        funcao = GRAFICOS_INTERATIVOS[grafico]
        resultado = funcao(**argumentos)

        return JSONResponse(resultado)

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        return JSONResponse({
            "erro": "Erro interno ao gerar gráfico interativo.",
            "detalhe": str(e),
            "traceback": tb
        }, status_code=500)
 



